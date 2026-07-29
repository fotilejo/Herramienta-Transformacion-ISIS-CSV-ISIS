# -*- coding: utf-8 -*-
"""
Generador de ISO 2709 a partir del CSV editado.

Paso 2 del ciclo:  CAT (ISIS) -> CSV -> edicion manual -> CAT (ISIS)

Compara el CSV editado contra la linea base, valida, informa que cambia y
genera el archivo ISO 2709 para importar en ABCD.

Modos:
  python generar_iso.py
      Solo los registros modificados. Es el modo normal de trabajo.

  python generar_iso.py --completo
      Los 49.592 registros. Sirve para la prueba de fidelidad: se importa
      en una base clonada VACIA, se reexporta con exportar_cat.py y se
      compara. Ver PROCEDIMIENTO.md.

  python generar_iso.py --completo --limite 2000
      Solo los primeros 2000 registros. El ISO completo pesa 41 MB y la subida
      web de ABCD suele tener un tope bastante mas bajo; para la prueba de
      fidelidad no hace falta la base entera.

  python generar_iso.py --mfns 023487,045552,045886
      Solo esos MFN. Sirve para armar un ISO chico con los registros dificiles
      (CP850, saltos de linea embebidos, subcampos) y probar con eso.

  python generar_iso.py --sin-mfn
      No agrega el campo con el MFN (ver INCLUIR_MFN abajo).

Nada de esto toca la base. La salida es un archivo .iso que se importa aparte.
"""

import csv
import os
import sys
from collections import Counter
from datetime import datetime

from openpyxl import load_workbook

csv.field_size_limit(10 ** 7)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
CSV_DIR  = os.path.join(BASE_DIR, "Salidas_CSV")
ISO_DIR  = os.path.join(BASE_DIR, "Salidas_ISO")

SEP_COLUMNAS    = ";"
SEP_OCURRENCIAS = "<|>"

# Caracteres de control dentro del contenido. Mismos tokens que exportar_cat.py.
# Se reponen en orden inverso al de la exportacion para que el byte vuelva igual.
CONTROLES = [("\r\n", "<NL>"), ("\r", "<CR>"), ("\n", "<LF>")]


# WXIS no puede importar un ISO que traiga CR o LF crudos dentro de un campo:
# corta el registro por la mitad y aborta con "execution error|import".
# Comprobado sobre ABCD: de 10 registros entraron 4 y el quinto rompio.
#
# Por eso, al escribir el ISO esos caracteres se ELIMINAN. En CAT son 684
# ocurrencias de basura de carga (un CR pegado al final de un descriptor,
# 'Quimica\r'), asi que sacarlos es una correccion. Igual queda registrado en
# el informe, porque es un cambio que no pediste en el CSV.
LIMPIADOS = []


def reponer_controles(texto, mfn=None, col=None):
    """Devuelve el texto con los tokens resueltos y sin caracteres de control."""
    for crudo, token in CONTROLES:
        texto = texto.replace(token, crudo)
    if "\r" in texto or "\n" in texto:
        limpio = texto.replace("\r\n", "").replace("\r", "").replace("\n", "")
        if mfn is not None:
            LIMPIADOS.append((mfn, col, texto, limpio))
        return limpio
    return texto

# ISO 2709 no tiene lugar para el numero de registro: al importar, CISIS asigna
# MFN nuevos. Para poder actualizar por MFN el numero tiene que viajar dentro
# del registro, como un campo mas. v900 no esta en la FDT de CAT, asi que no
# pisa nada. ISIS no valida la FDT, de modo que lo acepta sin problema.
INCLUIR_MFN = True
TAG_MFN     = "900"

# v401 y v110 son campos de sistema: ABCD/WXIS les agrega su propio sello cada
# vez que un registro se importa o se guarda (v401='true' marca que paso por
# importacion ISO; v110 guarda usuario+fecha+hash de la ultima edicion). No
# estan en la FDT de CAT, igual que v900.
#
# armar_registro() arma el registro de reemplazo completo, copiando de la
# linea base los campos que no se editaron -- inclusive estos dos. Si se
# manda el valor viejo, ABCD agrega su sello nuevo ENCIMA del que ya venia y
# quedan duplicados (comprobado: 121 celdas se separan en 128 - 120 esperadas
# de la correccion de Martinez Krahmer <|> 1 caracter de control limpiado
# <|> 6 casos de v401 duplicado <|> 1 caso de v110 duplicado, los 7 con
# patron 'valor' -> 'valor<|>valor'). Mismo mecanismo que el bug de v900:
# se soluciona igual, sin volver a mandar el valor viejo.
TAGS_SISTEMA = {"401", "110"}

# Separadores. El estandar ISO 2709 define 0x1E para fin de campo y 0x1D para
# fin de registro, pero CISIS no los usa: escribe '#' (0x23) para los dos.
# Verificado sobre un ISO real de mx de 43 MB: cero bytes 0x1D, 0x1E y 0x1F.
#
# No genera ambiguedad aunque el dato contenga '#', porque el registro no se
# parsea buscando separadores: el directorio da la posicion y el largo exactos
# de cada campo. El '#' es solo una marca visual.
FT = b"#"      # fin de campo
RT = b"#"      # fin de registro

# CISIS no escribe el ISO como un flujo continuo de bytes: parte cada registro
# en lineas de ANCHO_LINEA caracteres separadas por CRLF, y cierra tambien la
# ultima linea parcial. Verificado contra un ISO generado por el propio mx:
# el registro 1 mide 585 bytes = 7 lineas de 80 + una de 25, con CRLF en las 8.
#
# Sin este envoltorio el importador de ABCD lee un unico registro y se detiene.
ANCHO_LINEA = 80
FIN_LINEA   = b"\r\n"

# Limites del formato, verificados contra CAT (holgados: 11% y 43% de uso)
MAX_REGISTRO = 99999   # el largo va en 5 digitos
MAX_CAMPO    = 9999    # el largo de cada campo va en 4 digitos

# Orden en que se escriben los campos dentro del registro. mx los emite en el
# orden de la FDT (v3, v4, v2, v5, ...), que es el orden fisico en el .mst y
# tambien el de las columnas del CSV. Se completa al leer el encabezado.
ORDEN_TAGS = []


# =========================================================
# LECTURA DEL CSV
# =========================================================
def leer_csv(ruta):
    """Devuelve ({mfn: {tag: [ocurrencias]}}, orden, columnas)."""
    if not os.path.exists(ruta):
        sys.exit("ERROR: no se encontro %s" % ruta)

    registros, orden = {}, []
    with open(ruta, encoding="utf-8-sig", newline="") as f:
        r = csv.reader(f, delimiter=SEP_COLUMNAS)
        encabezados = next(r)
        if encabezados[0] != "MFN":
            sys.exit("ERROR: %s no tiene el formato esperado (falta la columna MFN)." % ruta)
        columnas = [(c.split("_")[0][1:], c) for c in encabezados[1:]]

        for nro, fila in enumerate(r, 2):
            if len(fila) != len(encabezados):
                sys.exit("ERROR: la fila %d de %s tiene %d columnas y deberia tener %d.\n"
                         "Puede que Excel haya guardado con otro separador."
                         % (nro, os.path.basename(ruta), len(fila), len(encabezados)))
            # Excel trata la columna MFN como numero y le saca los ceros de la
            # izquierda: '000001' vuelve como '1'. Se repone el relleno para que
            # el CSV editado siga emparejando contra la linea base.
            mfn = fila[0].strip()
            if mfn.isdigit():
                mfn = mfn.zfill(6)
            campos = {}
            for (tag, _), celda in zip(columnas, fila[1:]):
                if celda:
                    campos[tag] = celda.split(SEP_OCURRENCIAS)
            registros[mfn] = campos
            orden.append(mfn)
    return registros, orden, columnas


def buscar_pareja():
    """Ubica el par base/editable mas reciente en Salidas_CSV.

    El editable puede ser .xlsx (formato actual, celdas bloqueadas en Texto)
    o .csv (exportaciones viejas, antes del cambio a Excel). Se prefiere el
    .xlsx si estan los dos.
    """
    if not os.path.isdir(CSV_DIR):
        sys.exit("ERROR: no existe %s. Correr antes exportar_cat.py." % CSV_DIR)
    bases = sorted(f for f in os.listdir(CSV_DIR) if f.endswith("_base.csv"))
    if not bases:
        sys.exit("ERROR: no hay ningun *_base.csv en %s. Correr antes exportar_cat.py." % CSV_DIR)
    base = bases[-1]
    prefijo = base[:-len("_base.csv")]
    candidato_xlsx = prefijo + "_editable.xlsx"
    candidato_csv  = prefijo + "_editable.csv"
    if os.path.exists(os.path.join(CSV_DIR, candidato_xlsx)):
        editable = candidato_xlsx
    elif os.path.exists(os.path.join(CSV_DIR, candidato_csv)):
        editable = candidato_csv
    else:
        sys.exit("ERROR: existe %s pero falta %s (o su version .xlsx)." % (base, candidato_csv))
    return os.path.join(CSV_DIR, base), os.path.join(CSV_DIR, editable)


def leer_xlsx(ruta):
    """Como leer_csv, pero para el editable .xlsx. Mismo formato de salida."""
    if not os.path.exists(ruta):
        sys.exit("ERROR: no se encontro %s" % ruta)

    wb = load_workbook(ruta, read_only=True, data_only=True)
    ws = wb.active
    filas_iter = ws.iter_rows(values_only=True)
    encabezados = next(filas_iter, None)
    if not encabezados or encabezados[0] != "MFN":
        sys.exit("ERROR: %s no tiene el formato esperado (falta la columna MFN)." % ruta)
    columnas = [(c.split("_")[0][1:], c) for c in encabezados[1:]]

    registros, orden = {}, []
    for nro, fila in enumerate(filas_iter, 2):
        if fila is None or all(v is None for v in fila):
            continue  # fila vacia al final de la hoja
        valores = ["" if v is None else str(v) for v in fila]
        if len(valores) != len(encabezados):
            sys.exit("ERROR: la fila %d de %s tiene %d columnas y deberia tener %d."
                     % (nro, os.path.basename(ruta), len(valores), len(encabezados)))
        mfn = valores[0].strip()
        if mfn.isdigit():
            mfn = mfn.zfill(6)
        campos = {}
        for (tag, _), celda in zip(columnas, valores[1:]):
            if celda:
                campos[tag] = celda.split(SEP_OCURRENCIAS)
        registros[mfn] = campos
        orden.append(mfn)
    return registros, orden, columnas


def leer_editable(ruta):
    """Despacha a la lectura de .xlsx o .csv segun la extension del editable."""
    if ruta.lower().endswith(".xlsx"):
        return leer_xlsx(ruta)
    return leer_csv(ruta)


# =========================================================
# DIFERENCIAS
# =========================================================
def diferencias(base, editado, columnas, solo=None):
    """Devuelve (cambios, errores). cambios = {mfn: [(tag, col, viejo, nuevo)]}

    Si 'solo' trae una lista de tags, se comparan unicamente esos campos: el
    archivo editado es una proyeccion con pocas columnas y el resto no viaja en
    el, asi que su ausencia no significa que se hayan borrado.
    """
    nombre_col = dict(columnas)
    cambios, errores = {}, []

    faltantes = set(base) - set(editado)
    nuevos    = set(editado) - set(base)
    if faltantes:
        errores.append("Faltan %d MFN que estaban en la linea base (no se pueden borrar "
                       "registros desde el CSV): %s"
                       % (len(faltantes), ", ".join(sorted(faltantes)[:10])))
    if nuevos:
        errores.append("Hay %d MFN que no estaban en la linea base (no se pueden crear "
                       "registros desde el CSV): %s"
                       % (len(nuevos), ", ".join(sorted(nuevos)[:10])))

    for mfn in editado:
        if mfn not in base:
            continue
        candidatos = set(base[mfn]) | set(editado[mfn])
        if solo is not None:
            candidatos &= set(solo)
        for tag in sorted(candidatos, key=int):
            viejo = base[mfn].get(tag, [])
            nuevo = editado[mfn].get(tag, [])
            if viejo != nuevo:
                cambios.setdefault(mfn, []).append(
                    (tag, nombre_col.get(tag, "v" + tag),
                     SEP_OCURRENCIAS.join(viejo), SEP_OCURRENCIAS.join(nuevo)))

    # v9 lo asigna ABCD solo: no se puede editar desde el CSV
    for mfn, lista in cambios.items():
        for tag, col, viejo, nuevo in lista:
            if tag == "9":
                errores.append("MFN %s: v9 (numero de control) fue modificado de %r a %r. "
                               "Lo asigna ABCD automaticamente y no debe editarse."
                               % (mfn, viejo, nuevo))
    return cambios, errores


def validar_contenido(cambios):
    """Chequea lo que impediria escribir el ISO. Devuelve (errores, avisos)."""
    errores, avisos = [], []
    for mfn, lista in sorted(cambios.items()):
        for tag, col, viejo, nuevo in lista:
            for oc in nuevo.split(SEP_OCURRENCIAS):
                if not oc:
                    continue
                texto = reponer_controles(oc)
                try:
                    b = texto.encode("cp1252")
                except UnicodeEncodeError as e:
                    ch = texto[e.start]
                    errores.append(
                        "MFN %s %s: el caracter %r (U+%04X) no existe en cp1252 y la "
                        "base no puede guardarlo. Reemplazarlo por un equivalente. "
                        "Casos tipicos: flechas, simbolos matematicos, vinetas, letras "
                        "griegas o cirilicas, y espacios especiales copiados de la web."
                        % (mfn, col, ch, ord(ch)))
                    continue
                if len(b) + 1 > MAX_CAMPO:
                    errores.append("MFN %s %s: la ocurrencia ocupa %d bytes y el maximo "
                                   "por campo es %d." % (mfn, col, len(b) + 1, MAX_CAMPO))
                if ";" in oc:
                    avisos.append("MFN %s %s: contiene ';'. En ABCD el ';' separa "
                                  "ocurrencias en pantalla, pero aca es texto literal. "
                                  "Para separar ocurrencias va %s."
                                  % (mfn, col, SEP_OCURRENCIAS))
    return errores, avisos


# =========================================================
# ISO 2709
# =========================================================
def armar_registro(mfn, campos):
    """Construye un registro ISO 2709. campos = {tag: [ocurrencias]}"""
    entradas = []
    if INCLUIR_MFN:
        entradas.append((TAG_MFN, mfn))
    # En el orden de la FDT, igual que mx. Los tags que no figuren en el
    # encabezado del CSV se agregan al final, por numero, para no perderlos.
    # Se descarta cualquier "900" que venga en los datos (por ejemplo, si el
    # editable es de una base que ya paso antes por esta herramienta y quedo
    # ese campo sin limpiar): si no, queda duplicado -- el que se inyecta acá
    # mas el viejo -- y mx no sabe cual usar como MFN ('fatal: recupdat/mfn').
    #
    # Tambien se descartan v401 y v110 (ver TAGS_SISTEMA arriba): son campos
    # que ABCD sella solo en cada importacion. Si se manda el valor viejo,
    # ABCD agrega el nuevo encima y quedan duplicados.
    excluir = {TAG_MFN} | TAGS_SISTEMA
    orden = ([t for t in ORDEN_TAGS if t not in excluir] +
             sorted((t for t in campos if t not in ORDEN_TAGS and t not in excluir), key=int))
    for tag in orden:
        for oc in campos.get(tag, []):
            if oc:
                entradas.append((tag, oc))

    directorio, datos = b"", b""
    for tag, contenido in entradas:
        b = reponer_controles(contenido, mfn, "v%s" % tag).encode("cp1252") + FT
        directorio += ("%03d%04d%05d" % (int(tag), len(b), len(datos))).encode("ascii")
        datos += b

    base_datos = 24 + len(directorio) + 1
    largo      = base_datos + len(datos) + 1
    if largo > MAX_REGISTRO:
        sys.exit("ERROR: el registro MFN %s ocupa %d bytes y el maximo del formato "
                 "ISO 2709 es %d." % (mfn, largo, MAX_REGISTRO))

    # Leader de 24 caracteres, copiando exactamente lo que escribe CISIS.
    # Verificado contra un ISO real de mx: '00585' '00000' '0' '0' '00301' '000' '4500'.
    #   0-4   largo total del registro       5
    #   5-9   estado y codigos               5   ceros
    #   10    largo de indicadores           1   '0' (ISIS no usa indicadores)
    #   11    largo del codigo de subcampo   1   '0' (ISIS guarda ^x como texto)
    #   12-16 direccion base de los datos    5
    #   17-19 uso propio                     3   ceros
    #   20-23 mapa de entradas               4   '4500': largo 4, posicion 5
    lider = ("%05d" % largo + "00000" + "0" + "0" +
             "%05d" % base_datos + "000" + "4500")
    if len(lider) != 24:
        sys.exit("ERROR interno: el lider ISO 2709 quedo de %d caracteres "
                 "y debe ser de 24." % len(lider))

    registro = lider.encode("ascii") + directorio + FT + datos + RT
    return envolver(registro)


def envolver(registro):
    """Parte el registro en lineas de 80 con CRLF, como hace CISIS.

    La ultima linea, aunque sea parcial, tambien lleva CRLF. Sin esto el
    importador de ABCD lee un solo registro de todo el archivo.
    """
    return b"".join(registro[i:i + ANCHO_LINEA] + FIN_LINEA
                    for i in range(0, len(registro), ANCHO_LINEA))


# =========================================================
# MAIN
# =========================================================
def main():
    global INCLUIR_MFN
    completo = "--completo" in sys.argv
    if "--sin-mfn" in sys.argv:
        INCLUIR_MFN = False

    ruta_base, ruta_edit = buscar_pareja()
    print("Linea base : %s" % os.path.basename(ruta_base))
    print("Editado    : %s" % os.path.basename(ruta_edit))

    base, _, columnas       = leer_csv(ruta_base)
    editado, orden, cols_edit = leer_editable(ruta_edit)
    print("  %d y %d registros." % (len(base), len(editado)))

    ORDEN_TAGS[:] = [t for t, _ in columnas]

    # El editable puede ser una proyeccion: solo el MFN y las columnas elegidas.
    # En ese caso solo se comparan esas, y para armar el ISO se parte del
    # registro completo de la linea base y se le encima lo editado.
    tags_base = [t for t, _ in columnas]
    tags_edit = [t for t, _ in cols_edit]
    solo = None
    if set(tags_edit) < set(tags_base):
        solo = tags_edit
        print("  El editable trae solo %d columnas: %s"
              % (len(solo), ", ".join("v" + t for t in solo)))
        print("  El resto de los campos se toma intacto de la linea base.")
        completo_editado = {}
        for mfn, campos in base.items():
            fusion = dict(campos)
            for t in solo:
                if t in editado.get(mfn, {}):
                    fusion[t] = editado[mfn][t]
                else:
                    fusion.pop(t, None)
            completo_editado[mfn] = fusion
        editado = completo_editado

    print("\nComparando...")
    cambios, errores = diferencias(base, editado, columnas, solo)
    err2, avisos     = validar_contenido(cambios)
    errores += err2

    # Un porcentaje alto de registros modificados casi siempre significa que la
    # planilla reformateo una columna entera, no que se hayan editado tantos.
    if len(cambios) > len(base) * 0.05:
        por_campo = Counter(col for lista in cambios.values() for _, col, _, _ in lista)
        print("\n  ATENCION: cambiaron %d de %d registros (%.0f%%)."
              % (len(cambios), len(base), 100.0 * len(cambios) / len(base)))
        print("  Es mucho para una edicion manual. Suele pasar cuando Excel reformatea")
        print("  una columna entera (fechas, numeros con ceros a la izquierda).")
        print("  Campos con mas cambios:")
        for col, n in por_campo.most_common(6):
            print("     %-34s %d registros" % (col, n))
        print("  Revisar el informe antes de importar nada.")

    if errores:
        print("\nERROR: %d problemas impiden generar el ISO.\n" % len(errores))
        for e in errores[:30]:
            print("  " + e)
        if len(errores) > 30:
            print("  ... y %d mas" % (len(errores) - 30))
        sys.exit("\nGeneracion cancelada: no se escribio ningun archivo.")

    if avisos:
        print("\n  %d avisos (no bloquean):" % len(avisos))
        for a in avisos[:15]:
            print("    " + a)
        if len(avisos) > 15:
            print("    ... y %d mas" % (len(avisos) - 15))

    if "--mfns" in sys.argv:
        pedidos = sys.argv[sys.argv.index("--mfns") + 1].split(",")
        pedidos = [m.strip().zfill(6) for m in pedidos if m.strip()]
        faltan  = [m for m in pedidos if m not in editado]
        if faltan:
            sys.exit("ERROR: estos MFN no existen en el CSV: %s" % ", ".join(faltan))
        seleccion = [m for m in orden if m in set(pedidos)]
        print("\nModo MFN: se incluyen %d registros." % len(seleccion))
    elif completo:
        seleccion = orden
        print("\nModo COMPLETO: se incluyen los %d registros." % len(seleccion))
    else:
        seleccion = [m for m in orden if m in cambios]
        print("\nRegistros modificados: %d" % len(seleccion))
        if not seleccion:
            print("\nNo hay cambios. No se genera ningun archivo.")
            return

    if "--limite" in sys.argv:
        n = int(sys.argv[sys.argv.index("--limite") + 1])
        if n < len(seleccion):
            seleccion = seleccion[:n]
            print("  Limitado a los primeros %d." % n)

    os.makedirs(ISO_DIR, exist_ok=True)
    sello  = datetime.now().strftime("%Y%m%d_%H%M")
    sufijo = "completo" if completo else "cambios"
    if "--mfns" in sys.argv:
        sufijo = "dificiles"
    ruta_iso     = os.path.join(ISO_DIR, "cat_%s_%s.iso" % (sello, sufijo))
    ruta_informe = os.path.join(ISO_DIR, "cat_%s_%s.txt" % (sello, sufijo))

    with open(ruta_iso, "wb") as f:
        for mfn in seleccion:
            f.write(armar_registro(mfn, editado[mfn]))


    # Informe de cambios: se escribe siempre, aunque el modo sea completo.
    with open(ruta_informe, "w", encoding="utf-8") as f:
        f.write("Informe de cambios - %s\n" % datetime.now().strftime("%Y-%m-%d %H:%M"))
        f.write("Linea base: %s\nEditado   : %s\n" % (ruta_base, ruta_edit))
        f.write("Registros modificados: %d\n" % len(cambios))
        f.write("Registros en el ISO  : %d\n\n" % len(seleccion))
        for mfn in sorted(cambios):
            f.write("MFN %s\n" % mfn)
            for tag, col, viejo, nuevo in cambios[mfn]:
                f.write("   %s\n" % col)
                f.write("      antes : %s\n" % (viejo or "(vacio)"))
                f.write("      ahora : %s\n" % (nuevo or "(vacio)"))
            f.write("\n")
        if avisos:
            f.write("\nAvisos:\n")
            for a in avisos:
                f.write("  " + a + "\n")
        if LIMPIADOS:
            f.write("\n\nCaracteres de control eliminados (%d ocurrencias)\n" % len(LIMPIADOS))
            f.write("WXIS no puede importar un ISO con CR o LF crudos dentro de un campo.\n")
            f.write("Se quitaron al escribir el ISO. Es un cambio no pedido en el CSV:\n\n")
            for mfn, col, antes, despues in LIMPIADOS:
                f.write("  MFN %s %s\n" % (mfn, col))
                f.write("     antes : %r\n" % antes)
                f.write("     ahora : %r\n" % despues)

    tam = os.path.getsize(ruta_iso)
    print("\n--- Resumen ---")
    print("  Registros con cambios : %d" % len(cambios))
    print("  Registros en el ISO   : %d" % len(seleccion))
    print("  MFN en el campo v%s   : %s" % (TAG_MFN, "si" if INCLUIR_MFN else "no"))
    print("  Tamano del ISO        : %.1f MB" % (tam / 1048576))
    if LIMPIADOS:
        print("  Controles eliminados  : %d ocurrencias (CR/LF crudos; WXIS no los importa)"
              % len(LIMPIADOS))
        print("                          el detalle esta en el informe")
    print("\n  ISO     : %s" % ruta_iso)
    print("  Informe : %s" % ruta_informe)
    print("\nRevisar el informe antes de importar. Ver PROCEDIMIENTO.md.")


if __name__ == "__main__":
    main()
